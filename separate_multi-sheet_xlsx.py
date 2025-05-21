import openpyxl
from openpyxl import Workbook
import os
import sys

# Nome del file di partenza
file_input = sys.argv[1]
output_folder = "single_sheets"

# Crea la cartella di output se non esiste
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Carica il file Excel
wb = openpyxl.load_workbook(file_input)

# Per ogni foglio nel file originale
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]

    # Crea un nuovo workbook
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name

    # Copia i dati cella per cella
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            new_ws[cell.coordinate].value = cell.value

    # Salva il nuovo file
    output_file = os.path.join(output_folder, f"{sheet_name}.xlsx")
    new_wb.save(output_file)

print(f"Salvati {len(wb.sheetnames)} file in '{output_folder}'")
